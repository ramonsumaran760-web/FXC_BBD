"""
Services v2.0 — Orquestador de APIs de producción con Circuit Breakers
1. MarketDataService   → Yahoo Finance con CB
2. BrokerService       → Alpaca Paper Trading con CB
3. AMLService          → OpenSanctions + OFAC local con CB
4. RoboAdvisorService  → Claude API con CB
5. ReportService       → Excel multi-hoja + PDF ejecutivo
6. CacheService        → Redis con TTL, fallback en memoria
"""
import json, time, random, hashlib, os, logging
from datetime import datetime, timedelta, timezone
from typing import Optional
import requests

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────
# 7. REDIS CACHE — con fallback en memoria
# ─────────────────────────────────────────────────────────
_mem_cache: dict = {}

def cache_set(key: str, value, ttl: int = 60):
    try:
        import redis
        from core.config import settings
        r = redis.from_url(settings.REDIS_URL, decode_responses=True, socket_timeout=2)
        r.setex(key, ttl, json.dumps(value, default=str))
        return
    except Exception:
        pass
    _mem_cache[key] = {"v": value, "exp": time.time() + ttl}

def cache_get(key: str):
    try:
        import redis
        from core.config import settings
        r = redis.from_url(settings.REDIS_URL, decode_responses=True, socket_timeout=2)
        val = r.get(key)
        return json.loads(val) if val else None
    except Exception:
        pass
    entry = _mem_cache.get(key)
    if entry and entry["exp"] > time.time():
        return entry["v"]
    _mem_cache.pop(key, None)
    return None

def cache_delete(key: str):
    try:
        import redis
        from core.config import settings
        r = redis.from_url(settings.REDIS_URL, decode_responses=True, socket_timeout=2)
        r.delete(key)
    except Exception:
        pass
    _mem_cache.pop(key, None)

def cache_incr(key: str, ttl: int = 60) -> int:
    try:
        import redis
        from core.config import settings
        r = redis.from_url(settings.REDIS_URL, decode_responses=True, socket_timeout=2)
        pipe = r.pipeline()
        pipe.incr(key)
        pipe.expire(key, ttl)
        results = pipe.execute()
        return int(results[0])
    except Exception:
        entry = _mem_cache.get(key)
        val = (entry["v"] + 1) if entry and entry["exp"] > time.time() else 1
        _mem_cache[key] = {"v": val, "exp": time.time() + ttl}
        return val

# ─────────────────────────────────────────────────────────
# HELPERS — compartidos por market data y broker
# ─────────────────────────────────────────────────────────
ALPACA_DATA_URL = "https://data.alpaca.markets/v2"

def _alpaca_headers(key: str, secret: str) -> dict:
    return {"APCA-API-KEY-ID": key, "APCA-API-SECRET-KEY": secret,
            "Content-Type": "application/json"}

def _is_demo(key: str) -> bool:
    return not key or key in ("DEMO_KEY", "demo", "")


# ─────────────────────────────────────────────────────────
# 1. MARKET DATA — Alpaca Data API (precios reales)
# ─────────────────────────────────────────────────────────
TICKERS_DEFAULT = [
    # ── US Tecnología ──────────────────────────────────────
    "AAPL","MSFT","NVDA","TSLA","AMZN","GOOGL","META",
    "NFLX","AMD","INTC","CRM","ADBE","ORCL","QCOM","TXN",
    # ── US Finanzas ────────────────────────────────────────
    "JPM","BAC","GS","V","MA","WFC","AXP","C","MS","BLK",
    # ── US Salud ───────────────────────────────────────────
    "JNJ","UNH","PFE","ABBV","MRK","TMO","MRNA","AMGN",
    # ── US Energía ─────────────────────────────────────────
    "XOM","CVX","COP","SLB","OXY",
    # ── US Consumo ─────────────────────────────────────────
    "WMT","HD","MCD","KO","PEP","NKE","SBUX","DIS","COST",
    # ── US Industrial ──────────────────────────────────────
    "CAT","BA","GE","HON","RTX",
    # ── ETF Índices US ─────────────────────────────────────
    "SPY","QQQ","DIA","IWM","VTI",
    # ── ETF Sectores US ────────────────────────────────────
    "XLK","XLF","XLV","XLE","XLI","XLC",
    # ── ETF Internacional (otras bolsas) ───────────────────
    "EEM","EFA","EWJ","EWG","EWZ","FXI","IEFA","EWU","EWC",
    # ── ADR Globales (acciones internacionales en NYSE) ────
    "TSM","BABA","NVO","ASML","SAP","VALE","NIO","TM","SONY",
    # ── Renta Fija ─────────────────────────────────────────
    "BND","TLT","HYG","LQD","IEF",
    # ── Materias Primas ────────────────────────────────────
    "GLD","SLV","GDX","USO",
]
_price_cache: dict = {}
_price_ts: float = 0


def _alpaca_snapshot_bulk(tickers: list, key: str, secret: str) -> dict:
    """Obtiene precios reales de Alpaca Data API — endpoint /stocks/snapshots.
    Procesa en lotes de 30 para evitar límites de URL (~4 KB).
    """
    BATCH = 30
    ts_now = datetime.now(timezone.utc).isoformat()
    result = {}

    for i in range(0, len(tickers), BATCH):
        batch = tickers[i:i + BATCH]
        try:
            r = requests.get(
                f"{ALPACA_DATA_URL}/stocks/snapshots",
                params={"symbols": ",".join(batch), "feed": "iex"},
                headers=_alpaca_headers(key, secret),
                timeout=12,
            )
            if not r.ok:
                logger.warning(f"Alpaca snapshot batch {i//BATCH}: HTTP {r.status_code} {r.text[:200]}")
                continue
            raw = r.json()
        except Exception as e:
            logger.warning(f"Alpaca snapshot batch {i//BATCH}: {e}")
            continue

        for ticker in batch:
            snap  = raw.get(ticker) or {}
            daily = snap.get("dailyBar") or {}
            prev  = snap.get("prevDailyBar") or {}
            trade = snap.get("latestTrade") or {}
            price  = float(trade.get("p") or daily.get("c") or 0)
            open_  = float(daily.get("o") or price)
            prev_c = float(prev.get("c") or open_)
            chg_pct = round((price - prev_c) / prev_c * 100, 3) if prev_c else 0
            if price > 0:
                result[ticker] = {
                    "price":      round(price, 4),
                    "change_pct": chg_pct,
                    "open":       round(open_, 4),
                    "high":       round(float(daily.get("h") or price), 4),
                    "low":        round(float(daily.get("l") or price), 4),
                    "volume":     int(daily.get("v") or 0),
                    "source":     "alpaca_realtime",
                    "ts":         ts_now,
                }
    return result


def get_market_prices(tickers: list = None) -> dict:
    global _price_cache, _price_ts
    tickers = tickers or TICKERS_DEFAULT
    if time.time() - _price_ts < 15 and _price_cache:
        return {t: _price_cache[t] for t in tickers if t in _price_cache}

    from core.config import settings
    from core.circuit_breaker import cb_alpaca

    # ── Fuente primaria: Alpaca Data API ─────────────────
    if not _is_demo(settings.ALPACA_API_KEY):
        try:
            @cb_alpaca
            def _fetch_alpaca():
                return _alpaca_snapshot_bulk(tickers, settings.ALPACA_API_KEY,
                                             settings.ALPACA_API_SECRET)
            result = _fetch_alpaca()
            if result:
                _price_cache.update(result)
                _price_ts = time.time()
                return result
        except Exception as e:
            logger.warning(f"Alpaca Data API: {e}")

    # ── Fuente secundaria: Yahoo Finance ─────────────────
    try:
        from core.circuit_breaker import cb_yfinance
        @cb_yfinance
        def _fetch_yf():
            import yfinance as yf
            data = yf.download(tickers, period="1d", interval="1m",
                               progress=False, auto_adjust=True, threads=False)
            result = {}
            if not data.empty and "Close" in data.columns:
                for t in tickers:
                    try:
                        col = data["Close"][t] if len(tickers) > 1 else data["Close"]
                        col = col.dropna()
                        if len(col) >= 2:
                            close = float(col.iloc[-1])
                            open_ = float(col.iloc[0])
                            chg = round((close - open_) / open_ * 100, 3) if open_ else 0
                            result[t] = {"price": round(close, 4), "change_pct": chg,
                                         "open": round(open_, 4), "volume": 0,
                                         "source": "yahoo_finance",
                                         "ts": datetime.now(timezone.utc).isoformat()}
                    except Exception:
                        pass
            if not result:
                raise Exception("yfinance devolvió datos vacíos")
            return result
        result = _fetch_yf()
        _price_cache.update(result)
        _price_ts = time.time()
        return result
    except Exception as e:
        logger.warning(f"Yahoo Finance: {e}")

    # ── Fallback: última cache con datos reales ───────────
    if _price_cache:
        logger.warning("Sin datos frescos — devolviendo última cache de precios reales")
        return {t: _price_cache[t] for t in tickers if t in _price_cache}

    logger.error("Sin fuente de precios. Configure ALPACA_API_KEY y ALPACA_API_SECRET en Render.")
    return {}


def get_candles(ticker: str, period: str = "1mo", interval: str = "1d") -> list:
    from core.config import settings

    # ── Fuente primaria: Alpaca bars API ─────────────────
    if not _is_demo(settings.ALPACA_API_KEY):
        try:
            days = {"1d": 2, "5d": 5, "1mo": 30, "3mo": 90, "1y": 365, "2y": 730}.get(period, 30)
            tf   = {"1m": "1Min", "5m": "5Min", "15m": "15Min",
                    "1h": "1Hour", "1d": "1Day", "1wk": "1Week"}.get(interval, "1Day")
            start = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")
            r = requests.get(
                f"{ALPACA_DATA_URL}/stocks/{ticker}/bars",
                params={"timeframe": tf, "start": start, "feed": "iex", "limit": 1000},
                headers=_alpaca_headers(settings.ALPACA_API_KEY, settings.ALPACA_API_SECRET),
                timeout=12,
            )
            if r.ok:
                bars = r.json().get("bars") or []
                if bars:
                    return [
                        {"t": int(datetime.fromisoformat(b["t"].replace("Z", "+00:00")).timestamp() * 1000),
                         "o": round(b["o"], 4), "h": round(b["h"], 4),
                         "l": round(b["l"], 4), "c": round(b["c"], 4), "v": int(b["v"])}
                        for b in bars
                    ]
        except Exception as e:
            logger.warning(f"Alpaca bars {ticker}: {e}")

    # ── Fuente secundaria: Yahoo Finance ─────────────────
    try:
        import yfinance as yf
        df = yf.download(ticker, period=period, interval=interval,
                         progress=False, auto_adjust=True)
        if not df.empty:
            return [
                {"t": int(ts.timestamp() * 1000),
                 "o": round(float(row["Open"]), 4), "h": round(float(row["High"]), 4),
                 "l": round(float(row["Low"]), 4), "c": round(float(row["Close"]), 4),
                 "v": int(row["Volume"])}
                for ts, row in df.iterrows()
            ]
    except Exception as e:
        logger.warning(f"yfinance candles {ticker}: {e}")

    return []

# ─────────────────────────────────────────────────────────
# 2. BROKER — Alpaca Paper Trading con Circuit Breaker
# ─────────────────────────────────────────────────────────
ALPACA_BASE = "https://paper-api.alpaca.markets/v2"

def alpaca_get_account(key: str, secret: str) -> dict:
    if _is_demo(key):
        return {"equity": "12847.32", "cash": "3421.10", "buying_power": "6842.20",
                "portfolio_value": "12847.32", "status": "ACTIVE",
                "daytrade_count": 0, "source": "demo"}
    from core.circuit_breaker import cb_alpaca
    try:
        @cb_alpaca
        def _req():
            r = requests.get(f"{ALPACA_BASE}/account",
                             headers=_alpaca_headers(key, secret), timeout=8)
            return r.json() if r.ok else {"error": r.text}
        return _req()
    except Exception as e:
        return {"error": str(e)}

def alpaca_place_order(key: str, secret: str, ticker: str, monto_usd: float,
                       side: str = "buy", tipo: str = "market",
                       limit_price: float = None) -> dict:
    prices = get_market_prices([ticker])
    price = prices.get(ticker, {}).get("price", 100)
    fracciones = round(monto_usd / price, 8) if price > 0 else 0

    if _is_demo(key):
        return {"id": f"ord_{hashlib.md5(f'{ticker}{time.time()}'.encode()).hexdigest()[:12]}",
                "symbol": ticker, "side": side, "type": tipo,
                "notional": str(monto_usd), "filled_notional": str(monto_usd),
                "filled_qty": str(fracciones), "filled_avg_price": str(round(price, 4)),
                "status": "filled", "broker": "alpaca_paper_demo",
                "submitted_at": datetime.now(timezone.utc).isoformat(),
                "filled_at": datetime.now(timezone.utc).isoformat()}

    from core.circuit_breaker import cb_alpaca
    try:
        @cb_alpaca
        def _req():
            payload = {"symbol": ticker, "side": side, "type": tipo,
                       "time_in_force": "day" if tipo == "market" else "gtc",
                       "notional": str(round(monto_usd, 2))}
            if tipo == "limit" and limit_price:
                payload["limit_price"] = str(limit_price)
            r = requests.post(f"{ALPACA_BASE}/orders",
                              headers=_alpaca_headers(key, secret),
                              json=payload, timeout=12)
            return r.json() if r.ok else {"error": r.text}
        return _req()
    except Exception as e:
        return {"error": str(e)}

def alpaca_get_positions(key: str, secret: str) -> list:
    if _is_demo(key):
        return [
            {"symbol": "AAPL", "qty": "2.3456", "avg_entry_price": "189.50",
             "current_price": "195.20", "market_value": "457.89",
             "unrealized_pl": "13.38", "unrealized_plpc": "0.030"},
            {"symbol": "NVDA", "qty": "0.8721", "avg_entry_price": "850.00",
             "current_price": "912.40", "market_value": "795.27",
             "unrealized_pl": "54.40", "unrealized_plpc": "0.073"},
        ]
    from core.circuit_breaker import cb_alpaca
    try:
        @cb_alpaca
        def _req():
            r = requests.get(f"{ALPACA_BASE}/positions",
                             headers=_alpaca_headers(key, secret), timeout=8)
            return r.json() if r.ok else []
        return _req()
    except Exception:
        return []

def alpaca_cancel_order(key: str, secret: str, order_id: str) -> dict:
    if _is_demo(key):
        return {"status": "cancelled", "id": order_id}
    from core.circuit_breaker import cb_alpaca
    try:
        @cb_alpaca
        def _req():
            r = requests.delete(f"{ALPACA_BASE}/orders/{order_id}",
                                headers=_alpaca_headers(key, secret), timeout=8)
            return {"status": "cancelled"} if r.status_code == 204 else {"error": r.text}
        return _req()
    except Exception as e:
        return {"error": str(e)}

# ─────────────────────────────────────────────────────────
# 3. AML — OpenSanctions + OFAC local con Circuit Breaker
# ─────────────────────────────────────────────────────────
_LISTAS_NEGRAS_LOCAL = [
    "maduro", "al-qaeda", "isis", "farc", "hezbollah", "iran corp",
    "cartago cartel", "sinaloa", "taliban", "kim jong", "wagner group"
]

def aml_check_entidad(nombre: str, nit: str = None) -> dict:
    nombre_lower = nombre.lower()
    for bad in _LISTAS_NEGRAS_LOCAL:
        if bad in nombre_lower:
            return {"status": "blocked", "score": 1.0, "fuente": "local_ofac",
                    "detalle": f"Coincide con entidad sancionada: '{bad}'"}

    from core.circuit_breaker import cb_opensanctions
    try:
        @cb_opensanctions
        def _req():
            from core.config import settings
            payload = {"queries": {"q1": {"schema": "Company", "properties": {"name": [nombre]}}}}
            r = requests.post(settings.OPENSANCTIONS_URL,
                              json=payload,
                              headers={"Content-Type": "application/json"}, timeout=6)
            if r.ok:
                results = r.json().get("responses", {}).get("q1", {}).get("results", [])
                if results and results[0].get("score", 0) > 0.75:
                    return {"status": "alert", "score": results[0]["score"],
                            "fuente": "opensanctions",
                            "detalle": f"Posible coincidencia: {results[0].get('caption', '')}"}
                return {"status": "clear", "score": 0, "fuente": "opensanctions",
                        "detalle": "Sin coincidencias internacionales"}
            return None
        result = _req()
        if result:
            return result
    except Exception as e:
        logger.warning(f"OpenSanctions error: {e}")

    return {"status": "clear", "score": 0, "fuente": "local_fallback",
            "detalle": "Sin coincidencias detectadas (fallback local)"}

# ─────────────────────────────────────────────────────────
# 4. ECDSA — re-exportado desde security
# ─────────────────────────────────────────────────────────
from core.security import firmar_orden, verificar_firma, generar_nonce

# ─────────────────────────────────────────────────────────
# 5. ROBO-ADVISOR — Claude API con Circuit Breaker
# ─────────────────────────────────────────────────────────
ROBO_SYSTEM_PROMPT = """Eres un asesor de inversiones automatizado (Robo-Advisor) de InvestIQ.
Analiza el perfil del usuario y su portafolio. Responde EXCLUSIVAMENTE con JSON válido.
Estructura exacta requerida:
{
  "perfil": "conservador|moderado|agresivo",
  "score_riesgo": <entero 0-100>,
  "alerta_riesgo": <boolean>,
  "concentracion_max_ticker": "<TICKER>",
  "concentracion_max_pct": <float>,
  "sugerencia_rebalanceo": "<máximo 2 oraciones>",
  "acciones_recomendadas": ["<ticker - descripcion>", ...],
  "explicacion_voz": "<texto corto para TTS, máximo 30 palabras>"
}
No incluyas ningún texto fuera del JSON."""

def robo_advisor_analizar(perfil_usuario: dict, portafolio: list, claude_api_key: str = "") -> dict:
    from core.config import settings as cfg
    prompt_input = {
        "usuario_id": perfil_usuario.get("id", "usr_demo"),
        "edad": perfil_usuario.get("edad", 30),
        "ingresos_anuales_usd": perfil_usuario.get("ingresos_anuales_usd", 15000),
        "tolerancia_perdida": perfil_usuario.get("tolerancia_riesgo", "media"),
        "saldo_disponible_usd": perfil_usuario.get("saldo_usd", 0),
        "distribucion_portafolio": {
            p["ticker"]: f"{round(p['valor_total_usd'] / max(sum(x['valor_total_usd'] for x in portafolio), 1) * 100, 1)}%"
            for p in portafolio
        } if portafolio else {}
    }

    if claude_api_key and claude_api_key not in ("", "demo"):
        from core.circuit_breaker import cb_claude
        try:
            @cb_claude
            def _req():
                r = requests.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": claude_api_key,
                             "anthropic-version": "2023-06-01",
                             "Content-Type": "application/json"},
                    json={"model": cfg.CLAUDE_MODEL, "max_tokens": cfg.CLAUDE_MAX_TOKENS,
                          "system": ROBO_SYSTEM_PROMPT,
                          "messages": [{"role": "user", "content":
                              f"Analiza:\n{json.dumps(prompt_input, ensure_ascii=False, indent=2)}"}]},
                    timeout=20
                )
                if r.ok:
                    text = r.json()["content"][0]["text"].strip()
                    text = text.replace("```json", "").replace("```", "").strip()
                    resultado = json.loads(text)
                    resultado["_prompt_json_enviado"] = prompt_input
                    resultado["_modelo"] = cfg.CLAUDE_MODEL
                    return resultado
                raise Exception(f"Claude API error: {r.status_code}")
            return _req()
        except Exception as e:
            logger.warning(f"Claude API error: {e} — usando lógica local")

    # ── Lógica determinística local ──
    total = sum(p.get("valor_total_usd", 0) for p in portafolio) or 1
    concentracion = {p["ticker"]: round(p["valor_total_usd"] / total * 100, 2) for p in portafolio}
    max_ticker = max(concentracion, key=concentracion.get) if concentracion else "N/A"
    max_pct = concentracion.get(max_ticker, 0)
    edad = perfil_usuario.get("edad", 30)
    tolerancia = perfil_usuario.get("tolerancia_riesgo", "media")
    perfil = ("conservador" if tolerancia == "baja" or edad > 55
              else "agresivo" if tolerancia == "alta" and edad < 35 else "moderado")
    alerta = max_pct > 35 or (perfil == "conservador" and any(
        t in ["TSLA", "NVDA", "COIN"] for t in concentracion))
    return {
        "perfil": perfil,
        "score_riesgo": {"conservador": 30, "moderado": 55, "agresivo": 82}.get(perfil, 55),
        "alerta_riesgo": alerta,
        "concentracion_max_ticker": max_ticker,
        "concentracion_max_pct": max_pct,
        "sugerencia_rebalanceo": (
            f"{'Reducir ' + max_ticker + ' del ' + str(max_pct) + '% al máximo 25%. ' if alerta else 'Portafolio bien distribuido. '}"
            "Considerar BND o VTI para diversificación."
        ),
        "acciones_recomendadas": (
            ["BND - ETF Bonos del Tesoro USA", "VTI - ETF Mercado Total USA"] if perfil == "conservador"
            else ["QQQ - ETF Nasdaq 100", "NVDA - Semiconductores IA"] if perfil == "agresivo"
            else ["SPY - ETF S&P 500", "VTI - Mercado Total", "BND - Bonos 30%"]
        ),
        "explicacion_voz": f"Tu perfil es {perfil}. {'Alerta: concentración en ' + max_ticker + '.' if alerta else 'Portafolio equilibrado.'}",
        "_prompt_json_enviado": prompt_input,
        "_modelo": "local_deterministic"
    }

# ─────────────────────────────────────────────────────────
# 6. REPORTES — Excel multi-hoja + PDF ejecutivo
# ─────────────────────────────────────────────────────────
def generar_excel_portafolio(usuario: dict, portafolio: list, ordenes: list,
                              dividendos: list, ruta: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    VRD = PatternFill("solid", fgColor="0D7A4E")
    AZL = PatternFill("solid", fgColor="0D3D7A")
    WF  = Font(color="FFFFFF", bold=True, size=11)
    TF  = Font(bold=True, size=14, color="0D7A4E")
    borde = Border(
        left=Side(style="thin", color="2A3545"),
        right=Side(style="thin", color="2A3545"),
        top=Side(style="thin", color="2A3545"),
        bottom=Side(style="thin", color="2A3545")
    )

    def hoja_tabla(nombre, data, headers, fill):
        ws = wb.create_sheet(nombre)
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"] = f"InvestIQ — {nombre} — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A1"].font = TF; ws["A1"].alignment = Alignment(horizontal="center")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = fill; c.font = WF; c.alignment = Alignment(horizontal="center"); c.border = borde
        for ri, row in enumerate(data, 3):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = borde
                if ri % 2 == 0: c.fill = PatternFill("solid", fgColor="F0F7F4")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 45)

    hoja_tabla("Portafolio", [
        [p["ticker"], p.get("nombre",""), round(p.get("acciones",0),8),
         round(p.get("precio_promedio_compra",0),4), round(p.get("precio_actual",0),4),
         round(p.get("valor_total_usd",0),2), round(p.get("ganancia_perdida_usd",0),2),
         f"{round(p.get('ganancia_perdida_pct',0),4)}%"]
        for p in portafolio
    ], ["Ticker","Nombre","Acciones","Precio Compra","Precio Actual","Valor USD","G/P USD","G/P %"], VRD)

    hoja_tabla("Órdenes", [
        [o.get("id"), o.get("ticker"), o.get("tipo","").upper(),
         round(o.get("monto_usd",0),2), round(o.get("acciones",0),8),
         round(o.get("precio_ejecucion",0),4), o.get("estado"),
         "✓" if o.get("firma_verificada") else "✗", o.get("creado","")]
        for o in ordenes
    ], ["#","Ticker","Tipo","Monto USD","Acciones","Precio Ejec.","Estado","Firma","Fecha"], AZL)

    ws_res = wb.create_sheet("Resumen Fiscal")
    total_v = sum(p.get("valor_total_usd",0) for p in portafolio)
    total_gp = sum(p.get("ganancia_perdida_usd",0) for p in portafolio)
    total_div = sum(d.get("monto_usd",0) for d in dividendos)
    ws_res["A1"] = "RESUMEN FISCAL — INVESTIQ"
    ws_res["A1"].font = Font(bold=True, size=16, color="0D7A4E")
    ws_res.merge_cells("A1:C1")
    for ri, (a, b) in enumerate([
        ("Usuario", usuario.get("nombre","")),
        ("Email", usuario.get("email","")),
        ("Valor total portafolio", f"$ {total_v:,.2f}"),
        ("G/P capital", f"$ {total_gp:,.2f}"),
        ("Dividendos recibidos", f"$ {total_div:,.4f}"),
        ("Impuesto estimado (20%)", f"$ {max(total_gp,0)*0.20:,.2f}"),
        ("Generado", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ], 3):
        ws_res.cell(ri, 1, a).font = Font(bold=True)
        ws_res.cell(ri, 2, b)
        ws_res.column_dimensions["A"].width = 35
        ws_res.column_dimensions["B"].width = 30

    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    wb.save(ruta)
    return ruta

def generar_pdf_reporte(usuario: dict, portafolio: list, metricas: dict, ruta: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, HRFlowable, KeepTogether)
    from reportlab.lib.enums import TA_CENTER

    doc = SimpleDocTemplate(ruta, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    VRD = colors.HexColor("#0D7A4E"); AZL = colors.HexColor("#0D3D7A")
    LGRS = colors.HexColor("#F0F7F4")
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(
        f'<font color="#0D7A4E"><b>InvestIQ</b></font> — Reporte de Portafolio',
        ParagraphStyle("H", fontSize=20, alignment=TA_CENTER, spaceAfter=4)))
    story.append(Paragraph(
        f'Generado: {datetime.now().strftime("%d de %B de %Y %H:%M")} · {usuario.get("nombre","—")}',
        ParagraphStyle("S", fontSize=10, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=8)))
    story.append(HRFlowable(width="100%", color=VRD, thickness=2, spaceAfter=12))

    kpi_data = [
        ["Indicador", "Valor"],
        ["Valor total portafolio", f"$ {metricas.get('valor_portafolio',0):,.2f}"],
        ["Ganancia / Pérdida total", f"$ {metricas.get('ganancia_total',0):,.2f}"],
        ["Posiciones activas", str(metricas.get("posiciones",0))],
        ["Saldo disponible USD", f"$ {metricas.get('saldo_disponible',0):,.2f}"],
        ["Órdenes ejecutadas", str(metricas.get("ordenes_total",0))],
    ]
    t = Table(kpi_data, colWidths=[9*cm, 7*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),VRD), ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,-1),10),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LGRS]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("TOPPADDING",(0,0),(-1,-1),7), ("BOTTOMPADDING",(0,0),(-1,-1),7),
    ]))
    story.append(KeepTogether([Paragraph("<b>Indicadores Clave</b>", styles["Heading2"]),
                               Spacer(1,0.2*cm), t]))
    story.append(Spacer(1, 0.5*cm))

    port_data = [["Ticker","Acciones","Precio Compra","Precio Actual","Valor USD","G/P %"]]
    for p in portafolio:
        pct = p.get("ganancia_perdida_pct",0)
        port_data.append([p["ticker"], f"{p.get('acciones',0):.6f}",
                          f"$ {p.get('precio_promedio_compra',0):.4f}",
                          f"$ {p.get('precio_actual',0):.4f}",
                          f"$ {p.get('valor_total_usd',0):,.2f}",
                          f"{'+'if pct>=0 else ''}{pct:.2f}%"])
    tp = Table(port_data, colWidths=[2*cm, 3*cm, 3.5*cm, 3.5*cm, 3.5*cm, 2.5*cm])
    tp.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),AZL), ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"), ("FONTSIZE",(0,0),(-1,-1),9),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#EEF4FB")]),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#DDDDDD")),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))
    story.append(KeepTogether([Paragraph("<b>Posiciones del Portafolio</b>", styles["Heading2"]),
                               Spacer(1,0.2*cm), tp]))
    story.append(Spacer(1,0.4*cm))
    story.append(Paragraph(
        f'<i>InvestIQ · IA activo · Alpaca Paper Trading · {datetime.now().year}</i>',
        ParagraphStyle("F", fontSize=8, textColor=colors.grey, alignment=TA_CENTER)))
    doc.build(story)
    return ruta
